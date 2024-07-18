import streamlit as st
import openpyxl 
import io
from openpyxl.styles import PatternFill

# ------------------------------------------------------------------------------------------------------------------------------

def remove_URL_spaces_xlsx(doc, new_sheet="New URL sheet"):
    """
    This function allows to remove all possible spaces presents in URLs, it's worth noting that the space must be represented with the following symbol: %.
    
    Parameters:
        First, a document of type '.xlsx' with the links to verify.
        Second, the name of the sheet that you want to create to make the possible changes, it is not necessary to add any extension."    
    The function creates a second sheet with the desired name where, in the first column, it displays a message with the status of the original URL, and in the next column, either the original URL or the new one depending on the status of the original URL.
        If the original URL had any spaces, a message indicating this issue is returned along with the number of spaces it contained, and the same URL but with all spaces removed ready to be used! In addition fills in the message cell in red.
        If the URL does not contain any spaces, a message indicating this is returned along with the original URL. In addition fills in the message cell in green.
    Returns:
        The corresponding document. 
    """
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid") # Define the type of filling that is desired for a cell.
    green_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
    workbook = openpyxl.load_workbook(doc) # Open the Excel document.
    URL_file = workbook.worksheets[0] # Create new sheet.
    new_URL_sheet = workbook.create_sheet(title = new_sheet) # Add a name to the created sheet.
    for url_index in range(2, URL_file.max_row+1): # Iterate over the second column, and starting from the second row, we iterate over the URLs.
        url = URL_file.cell(row=url_index,column=1).value # Obtain the URL from the cell.
        if "%" in url: 
            new_URL = ""
            num_espaces = 0
            for letter in url:
                if letter != "%":
                    new_URL = new_URL + letter
                elif letter == "%":
                    num_espaces = num_espaces + 1
            message = ("There were " + str(num_espaces) + " spaces in your URL on line " + str(url_index+1)+ ", this is your new URL without spaces")
            new_URL_sheet.cell(row=url_index, column=1, value = message)
            new_URL_sheet.cell(row=url_index, column=1).fill = red_fill
            new_URL_sheet.cell(row=url_index, column=2, value = new_URL)
        else:
            message = ("There were no space in your URL on the line " + str(url_index+1))
            new_URL_sheet.cell(row=url_index, column=1, value = message)
            new_URL_sheet.cell(row=url_index, column=1).fill = green_fill
            new_URL_sheet.cell(row=url_index, column=2, value = url)
    workbook.save(doc)
    output = io.BytesIO()
    workbook.save(output)
    st.download_button(label = "📥 Download file with corrected URLs", data = output, file_name =  str(new_sheet) + ".xlsx", mime = "application/vnd.ms-excel")
